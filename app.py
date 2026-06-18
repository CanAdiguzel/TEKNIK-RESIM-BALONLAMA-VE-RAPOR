from __future__ import annotations

import io
import math
import os
import re
from datetime import date
from pathlib import Path
from typing import Any

import fitz
import pandas as pd
import streamlit as st
import streamlit.elements.image as streamlit_legacy_image
from streamlit.elements.lib.image_utils import image_to_url as current_image_to_url
from streamlit.elements.lib.layout_utils import LayoutConfig
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


# streamlit-drawable-canvas 0.9.3, Streamlit 1.58'de taşınan image_to_url
# fonksiyonunun eski yerini kullanır. Bu küçük uyumluluk katmanı mevcut API'ye yönlendirir.
if not hasattr(streamlit_legacy_image, "image_to_url"):
    def _legacy_image_to_url(image, width, clamp, channels, output_format, image_id):
        return current_image_to_url(
            image,
            LayoutConfig(width=width),
            clamp,
            channels,
            output_format,
            image_id,
        )

    streamlit_legacy_image.image_to_url = _legacy_image_to_url

from streamlit_drawable_canvas import st_canvas


APP_TITLE = "AS9102 Teknik Resim Balonlama"
RENDER_DPI = 120
RED = (220, 20, 45)

FORM3_COLUMNS = [
    "Characteristic No / Balon No",
    "Sheet",
    "Zone",
    "View",
    "Characteristic Type",
    "Characteristic Designator",
    "Drawing Requirement",
    "Nominal",
    "Upper Tolerance",
    "Lower Tolerance",
    "Upper Limit",
    "Lower Limit",
    "Units",
    "GD&T / Datum Reference",
    "Quantity",
    "Inspection Method",
    "Measuring Equipment",
    "Designed / Qualified Tooling",
    "Result 1",
    "Result 2",
    "Result 3",
    "Result Summary",
    "Acceptance Status",
    "Nonconformance No",
    "Inspector",
    "Inspection Date",
    "Remarks",
    "_target_x",
    "_target_y",
    "_balloon_x",
    "_balloon_y",
]

VISIBLE_COLUMNS = FORM3_COLUMNS[:-4]


def initialize_state() -> None:
    defaults = {
        "pdf_bytes": None,
        "pdf_name": None,
        "page_images": [],
        "page_sizes": [],
        "records": pd.DataFrame(columns=FORM3_COLUMNS),
        "canvas_version": 0,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


@st.cache_data(show_spinner=False)
def render_pdf(pdf_bytes: bytes, dpi: int = RENDER_DPI) -> tuple[list[bytes], list[dict[str, float]]]:
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    matrix = fitz.Matrix(dpi / 72, dpi / 72)
    images: list[bytes] = []
    sizes: list[dict[str, float]] = []
    try:
        for page in document:
            pixmap = page.get_pixmap(matrix=matrix, alpha=False)
            images.append(pixmap.tobytes("png"))
            sizes.append(
                {
                    "image_width": pixmap.width,
                    "image_height": pixmap.height,
                    "pdf_width": page.rect.width,
                    "pdf_height": page.rect.height,
                }
            )
    finally:
        document.close()
    return images, sizes


def as_number(value: Any) -> float | None:
    if value is None or pd.isna(value) or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_requirement(text: str) -> tuple[float | None, float | None, float | None]:
    normalized = text.replace(",", ".").replace("−", "-").upper()
    nominal_match = re.search(r"(?:Ø|⌀|R)?\s*(\d+(?:\.\d+)?)", normalized)
    nominal = float(nominal_match.group(1)) if nominal_match else None
    plus_minus = re.search(r"±\s*(\d+(?:\.\d+)?)", normalized)
    if plus_minus:
        tolerance = float(plus_minus.group(1))
        return nominal, tolerance, tolerance
    upper_match = re.search(r"\+\s*(\d+(?:\.\d+)?)", normalized)
    lower_match = re.search(r"-\s*(\d+(?:\.\d+)?)", normalized)
    return (
        nominal,
        float(upper_match.group(1)) if upper_match else None,
        float(lower_match.group(1)) if lower_match else None,
    )


def suggest_method(requirement: str, parameter: str, characteristic_type: str) -> tuple[str, str]:
    text = f"{requirement} {parameter} {characteristic_type}".casefold()
    if any(term in text for term in ("diş", "thread", " m", "m6", "m8", "m10", "m12", "go/no")):
        return "GO/NO-GO diş mastarı ile kontrol", "GO/NO-GO Diş Mastarı"
    if any(term in text for term in ("pozisyon", "position", "profil", "profile", "gd&t", "datum", "paralel", "diklik")):
        return "CMM ölçümü", "CMM"
    if any(term in text for term in ("pürüz", "roughness", "ra ")):
        return "Yüzey pürüzlülüğü ölçümü", "Profilometre"
    if any(term in text for term in ("h7", "iç çap", "delik", "bore")):
        return "İç çap kontrolü", "İç Çap Komparatörü / Tampon Mastar / CMM"
    if any(term in text for term in ("dış çap", "outer diameter", "ø", "⌀", "çap")):
        return "Dış çap ölçümü", "Mikrometre"
    if any(term in text for term in ("pah", "chamfer", "radyüs", "radius", " r")):
        return "Profil kontrolü", "Radyüs/Pah Mastarı veya Profil Projektör"
    if any(term in text for term in ("malzeme", "material", "proses", "process", "sertifika", "genel not")):
        return "Doküman doğrulama", "Sertifika / Proses Kayıt Kontrolü"
    return "Boyutsal ölçüm", "Kumpas / Mikrometre / Yükseklik Mihengiri"


def empty_record(number: int, page: int, target_x: float, target_y: float) -> dict[str, Any]:
    balloon_x = min(0.96, target_x + 0.08)
    balloon_y = max(0.04, target_y - 0.06)
    return {
        "Characteristic No / Balon No": number,
        "Sheet": page,
        "Zone": "",
        "View": "",
        "Characteristic Type": "Dimensional",
        "Characteristic Designator": "",
        "Drawing Requirement": "",
        "Nominal": None,
        "Upper Tolerance": None,
        "Lower Tolerance": None,
        "Upper Limit": None,
        "Lower Limit": None,
        "Units": "mm",
        "GD&T / Datum Reference": "",
        "Quantity": 1,
        "Inspection Method": "Boyutsal ölçüm",
        "Measuring Equipment": "Kumpas / Mikrometre / Yükseklik Mihengiri",
        "Designed / Qualified Tooling": "N/A",
        "Result 1": "Ölçüm bekleniyor",
        "Result 2": "Ölçüm bekleniyor",
        "Result 3": "Ölçüm bekleniyor",
        "Result Summary": "Ölçüm bekleniyor",
        "Acceptance Status": "Bekliyor",
        "Nonconformance No": "",
        "Inspector": "Doldurulacak",
        "Inspection Date": "Doldurulacak",
        "Remarks": "",
        "_target_x": target_x,
        "_target_y": target_y,
        "_balloon_x": balloon_x,
        "_balloon_y": balloon_y,
    }


def apply_suggestions(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for index, row in result.iterrows():
        requirement = str(row.get("Drawing Requirement") or "")
        nominal = as_number(row.get("Nominal"))
        upper = as_number(row.get("Upper Tolerance"))
        lower = as_number(row.get("Lower Tolerance"))
        if requirement and nominal is None:
            parsed_nominal, parsed_upper, parsed_lower = parse_requirement(requirement)
            nominal = parsed_nominal
            upper = upper if upper is not None else parsed_upper
            lower = lower if lower is not None else parsed_lower
            result.at[index, "Nominal"] = nominal
            result.at[index, "Upper Tolerance"] = upper
            result.at[index, "Lower Tolerance"] = lower
        if nominal is not None and upper is not None:
            result.at[index, "Upper Limit"] = nominal + upper
        if nominal is not None and lower is not None:
            result.at[index, "Lower Limit"] = nominal - abs(lower)
        method, equipment = suggest_method(
            requirement,
            str(row.get("Characteristic Designator") or ""),
            str(row.get("Characteristic Type") or ""),
        )
        if not str(row.get("Inspection Method") or "").strip() or row.get("Inspection Method") == "Boyutsal ölçüm":
            result.at[index, "Inspection Method"] = method
        if not str(row.get("Measuring Equipment") or "").strip() or row.get("Measuring Equipment") == "Kumpas / Mikrometre / Yükseklik Mihengiri":
            result.at[index, "Measuring Equipment"] = equipment
    return result


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        "C:/Windows/Fonts/arialbd.ttf",
        "C:/Windows/Fonts/calibrib.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def draw_preview(image_bytes: bytes, rows: pd.DataFrame, page_number: int) -> Image.Image:
    image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    draw = ImageDraw.Draw(image)
    font = load_font(max(18, round(image.width / 65)))
    radius = max(18, round(image.width / 70))
    for _, row in rows.iterrows():
        if int(as_number(row.get("Sheet")) or 0) != page_number:
            continue
        number = int(as_number(row.get("Characteristic No / Balon No")) or 0)
        target = (float(row["_target_x"]) * image.width, float(row["_target_y"]) * image.height)
        balloon = (float(row["_balloon_x"]) * image.width, float(row["_balloon_y"]) * image.height)
        if math.dist(target, balloon) > radius * 1.6:
            draw.line([balloon, target], fill=RED, width=max(2, radius // 7))
            draw.ellipse(
                [target[0] - 3, target[1] - 3, target[0] + 3, target[1] + 3],
                fill=RED,
            )
        draw.ellipse(
            [balloon[0] - radius, balloon[1] - radius, balloon[0] + radius, balloon[1] + radius],
            outline=RED,
            fill="white",
            width=max(3, radius // 6),
        )
        text = str(number)
        box = draw.textbbox((0, 0), text, font=font)
        draw.text(
            (balloon[0] - (box[2] - box[0]) / 2, balloon[1] - (box[3] - box[1]) / 2 - 2),
            text,
            fill=RED,
            font=font,
        )
    return image


def extract_canvas_points(canvas_json: dict[str, Any] | None, display_width: int, display_height: int) -> list[tuple[float, float]]:
    points: list[tuple[float, float]] = []
    for obj in (canvas_json or {}).get("objects", []):
        if obj.get("type") != "circle":
            continue
        left = float(obj.get("left", 0))
        top = float(obj.get("top", 0))
        radius = float(obj.get("radius", 0))
        center_x = left + radius
        center_y = top + radius
        points.append((center_x / display_width, center_y / display_height))
    return points


def create_simple_balloon_pdf(pdf_bytes: bytes, records: pd.DataFrame) -> bytes:
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    red = (0.86, 0.04, 0.12)
    try:
        for page_index, page in enumerate(document, start=1):
            page_rows = records[records["Sheet"].apply(lambda value: int(as_number(value) or 0) == page_index)]
            for _, row in page_rows.iterrows():
                number = int(as_number(row["Characteristic No / Balon No"]) or 0)
                target = fitz.Point(float(row["_target_x"]) * page.rect.width, float(row["_target_y"]) * page.rect.height)
                balloon = fitz.Point(float(row["_balloon_x"]) * page.rect.width, float(row["_balloon_y"]) * page.rect.height)
                radius = 11
                if math.dist((target.x, target.y), (balloon.x, balloon.y)) > radius * 1.7:
                    angle = math.atan2(target.y - balloon.y, target.x - balloon.x)
                    line_start = fitz.Point(
                        balloon.x + radius * math.cos(angle),
                        balloon.y + radius * math.sin(angle),
                    )
                    page.draw_line(line_start, target, color=red, width=1.4, overlay=True)
                    page.draw_circle(target, radius=2, color=red, fill=red, overlay=True)
                page.draw_circle(balloon, radius=radius, color=red, fill=(1, 1, 1), width=1.7, overlay=True)
                text = str(number)
                text_width = fitz.get_text_length(text, fontname="helv", fontsize=9)
                page.insert_text(
                    fitz.Point(balloon.x - text_width / 2, balloon.y + 3.2),
                    text,
                    fontname="helv",
                    fontsize=9,
                    color=red,
                    overlay=True,
                )
        output = io.BytesIO()
        document.save(output, garbage=4, deflate=True)
        return output.getvalue()
    finally:
        document.close()


def style_title(ws, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 15


def standard_border() -> Border:
    side = Side(style="thin", color="4F4F4F")
    return Border(left=side, right=side, top=side, bottom=side)


def write_kv_sheet(ws, title: str, fields: list[tuple[str, str]]) -> None:
    style_title(ws, title, 6)
    border = standard_border()
    for row_index, (label, value) in enumerate(fields, start=4):
        ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=3)
        ws.merge_cells(start_row=row_index, start_column=4, end_row=row_index, end_column=6)
        label_cell = ws.cell(row_index, 2, label)
        value_cell = ws.cell(row_index, 4, value)
        label_cell.font = Font(bold=True)
        label_cell.fill = PatternFill("solid", fgColor="D9E2F3")
        for column in range(2, 7):
            ws.cell(row_index, column).border = border
            ws.cell(row_index, column).alignment = Alignment(vertical="center", wrap_text=True)
        value_cell.fill = PatternFill("solid", fgColor="FFFFFF")
        ws.row_dimensions[row_index].height = 25
    for column, width in {"A": 3, "B": 20, "C": 12, "D": 24, "E": 18, "F": 18}.items():
        ws.column_dimensions[column].width = width
    ws.sheet_view.showGridLines = False


def create_as9102_workbook(
    records: pd.DataFrame,
    part_info: dict[str, str],
    process_info: dict[str, str],
) -> bytes:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "AS9102_Form3_Olcum_Raporu"
    style_title(ws, "AS9102 FAI FORM 3 - ÖLÇÜM RAPORU", len(VISIBLE_COLUMNS))

    header_row = 4
    border = standard_border()
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    for column_index, title in enumerate(VISIBLE_COLUMNS, start=1):
        cell = ws.cell(header_row, column_index, title)
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 52

    sorted_records = records.sort_values("Characteristic No / Balon No", kind="stable")
    for row_index, (_, record) in enumerate(sorted_records.iterrows(), start=5):
        for column_index, key in enumerate(VISIBLE_COLUMNS, start=1):
            value = record.get(key)
            if value is None or pd.isna(value):
                value = ""
            cell = ws.cell(row_index, column_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=9)
            if key == "Acceptance Status":
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            if key == "Characteristic No / Balon No":
                cell.font = Font(name="Arial", size=10, bold=True, color="C00000")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(VISIBLE_COLUMNS))}{max(5, ws.max_row)}"
    widths = {
        1: 14, 2: 8, 3: 9, 4: 12, 5: 18, 6: 20, 7: 25, 8: 11, 9: 12,
        10: 12, 11: 12, 12: 12, 13: 9, 14: 20, 15: 9, 16: 23, 17: 26,
        18: 22, 19: 18, 20: 18, 21: 18, 22: 18, 23: 15, 24: 18, 25: 16,
        26: 16, 27: 28,
    }
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"

    form1 = workbook.create_sheet("Form1_Parca_Bilgileri")
    write_kv_sheet(
        form1,
        "AS9102 FORM 1 - PARÇA BİLGİLERİ",
        [
            ("Parça Numarası", part_info.get("part_number", "")),
            ("Parça Adı", part_info.get("part_name", "")),
            ("Seri / Lot Numarası", part_info.get("serial_lot", "")),
            ("FAI Rapor Numarası", part_info.get("fai_report_no", "")),
            ("Çizim Numarası", part_info.get("drawing_number", "")),
            ("Çizim Revizyonu", part_info.get("drawing_revision", "")),
            ("Kuruluş / Tedarikçi", part_info.get("supplier", "")),
            ("Müşteri", part_info.get("customer", "")),
            ("Sipariş / İş Emri", part_info.get("order_no", "")),
            ("FAI Türü", part_info.get("fai_type", "Full FAI")),
        ],
    )

    form2 = workbook.create_sheet("Form2_Malzeme_Proses")
    write_kv_sheet(
        form2,
        "AS9102 FORM 2 - MALZEME VE PROSES",
        [
            ("Malzeme", process_info.get("material", "")),
            ("Malzeme Şartnamesi", process_info.get("material_spec", "")),
            ("Sertifika Numarası", process_info.get("certificate_no", "")),
            ("Özel Proses", process_info.get("special_process", "")),
            ("Proses Şartnamesi", process_info.get("process_spec", "")),
            ("Proses Tedarikçisi", process_info.get("process_supplier", "")),
            ("Fonksiyonel Test", process_info.get("functional_test", "")),
            ("Açıklama", process_info.get("remarks", "")),
        ],
    )

    methods = workbook.create_sheet("Olcum_Metodu_Listesi")
    style_title(methods, "ÖLÇÜM METODU VE EKİPMAN ÖNERİ LİSTESİ", 4)
    method_rows = [
        ("Karakteristik", "Önerilen Metot", "Ölçüm Ekipmanı", "Not"),
        ("Dış çap", "Dış çap ölçümü", "Mikrometre", "Çap ve tolerans aralığına uygun"),
        ("İç çap / H7 delik", "İç çap kontrolü", "İç çap komparatörü / Tampon mastar / CMM", ""),
        ("Diş", "GO/NO-GO kontrolü", "GO/NO-GO diş mastarı", ""),
        ("Doğrusal ölçü", "Boyutsal ölçüm", "Kumpas / Mikrometre / Yükseklik mihengiri", ""),
        ("Pozisyon / Profil / GD&T", "Koordinat ölçümü", "CMM", ""),
        ("Yüzey pürüzlülüğü", "Pürüzlülük ölçümü", "Profilometre", ""),
        ("Pah / Radyüs", "Profil kontrolü", "Mastar / Profil projektör", ""),
        ("Genel not / Malzeme / Proses", "Doküman doğrulama", "Sertifika / Proses kayıtları", ""),
    ]
    for row_index, row in enumerate(method_rows, start=4):
        for column_index, value in enumerate(row, start=1):
            cell = methods.cell(row_index, column_index, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_index == 4:
                cell.font = Font(bold=True)
                cell.fill = header_fill
    for index, width in enumerate((25, 27, 42, 30), start=1):
        methods.column_dimensions[get_column_letter(index)].width = width
    methods.freeze_panes = "A5"
    methods.auto_filter.ref = f"A4:D{methods.max_row}"
    methods.sheet_view.showGridLines = False

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def add_ocr_candidates(pdf_bytes: bytes, rotation: int) -> pd.DataFrame:
    try:
        from backend.ocr_utils import analyze_pdf_ocr
    except Exception:
        return pd.DataFrame(columns=FORM3_COLUMNS)

    temp_path = Path("work") / "streamlit_ocr_input.pdf"
    temp_path.parent.mkdir(exist_ok=True)
    temp_path.write_bytes(pdf_bytes)
    annotations = analyze_pdf_ocr(temp_path, rotation)
    rows: list[dict[str, Any]] = []
    for item in annotations:
        record = empty_record(
            item.balloon_no,
            item.page,
            item.target_bbox.x + item.target_bbox.width / 2,
            item.target_bbox.y + item.target_bbox.height / 2,
        )
        record.update(
            {
                "Characteristic Type": item.tolerance_type,
                "Characteristic Designator": item.parameter,
                "Drawing Requirement": item.dimension_text,
                "Nominal": item.nominal,
                "Upper Tolerance": item.upper_tol,
                "Lower Tolerance": item.lower_tol,
                "Upper Limit": item.max_limit,
                "Lower Limit": item.min_limit,
                "GD&T / Datum Reference": item.gdnt_reference,
                "Remarks": "Yerel OCR adayı; teknik olarak doğrulayın.",
            }
        )
        method, equipment = suggest_method(item.dimension_text, item.parameter, item.tolerance_type)
        record["Inspection Method"] = method
        record["Measuring Equipment"] = equipment
        rows.append(record)
    return pd.DataFrame(rows, columns=FORM3_COLUMNS)


def editor_configuration() -> dict[str, Any]:
    config: dict[str, Any] = {
        "Characteristic No / Balon No": st.column_config.NumberColumn("Characteristic No / Balon No", min_value=1, step=1, required=True),
        "Sheet": st.column_config.NumberColumn("Sheet", min_value=1, step=1),
        "Nominal": st.column_config.NumberColumn("Nominal", format="%.5f"),
        "Upper Tolerance": st.column_config.NumberColumn("Upper Tolerance", format="%.5f"),
        "Lower Tolerance": st.column_config.NumberColumn("Lower Tolerance", format="%.5f"),
        "Upper Limit": st.column_config.NumberColumn("Upper Limit", format="%.5f"),
        "Lower Limit": st.column_config.NumberColumn("Lower Limit", format="%.5f"),
        "Drawing Requirement": st.column_config.TextColumn("Drawing Requirement", width="large"),
        "Measuring Equipment": st.column_config.TextColumn("Measuring Equipment", width="large"),
        "Remarks": st.column_config.TextColumn("Remarks", width="large"),
    }
    for hidden in ("_target_x", "_target_y", "_balloon_x", "_balloon_y"):
        config[hidden] = None
    return config


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🔴", layout="wide")
    initialize_state()
    st.title("🔴 AS9102 Teknik Resim Balonlama")
    st.caption(
        "PDF üzerinde yalnız sade kırmızı numaralı balonlar oluşturur; tüm teknik bilgiler "
        "AS9102 FAI Form 3 mantığında Excel raporuna aktarılır."
    )

    with st.sidebar:
        st.header("Çalışma modu")
        st.success("Manuel mod her zaman kullanılabilir. OpenAI API zorunlu değildir.")
        rotation = st.selectbox(
            "OCR için sayfa döndürme",
            [0, 90, 180, 270],
            index=3,
            help="Taranmış çizim yan duruyorsa 270 genellikle doğru yöndür.",
        )
        st.info("SLDDRW desteklenmez. Lütfen PDF çıktısı yükleyin.")

    uploaded = st.file_uploader("PDF teknik resmi yükleyin", type=["pdf"])
    if uploaded is None:
        st.info("Başlamak için PDF dosyası yükleyin.")
        return

    pdf_bytes = uploaded.getvalue()
    if st.session_state.pdf_name != uploaded.name or st.session_state.pdf_bytes != pdf_bytes:
        try:
            images, sizes = render_pdf(pdf_bytes)
        except Exception as exc:
            st.error(f"PDF okunamadı: {exc}")
            return
        st.session_state.pdf_name = uploaded.name
        st.session_state.pdf_bytes = pdf_bytes
        st.session_state.page_images = images
        st.session_state.page_sizes = sizes
        st.session_state.records = pd.DataFrame(columns=FORM3_COLUMNS)
        st.session_state.canvas_version += 1

    st.success(f"{uploaded.name} yüklendi · {len(st.session_state.page_images)} sayfa")

    tool_col1, tool_col2, tool_col3 = st.columns([1, 1, 2])
    with tool_col1:
        if st.button("Yerel OCR ile aday bul", use_container_width=True):
            with st.spinner("Yerel OCR çalışıyor…"):
                try:
                    candidates = add_ocr_candidates(pdf_bytes, rotation)
                    st.session_state.records = apply_suggestions(candidates)
                    st.session_state.canvas_version += 1
                    if candidates.empty:
                        st.warning("OCR adayı bulunamadı. Manuel balon eklemeye devam edebilirsiniz.")
                    else:
                        st.success(f"{len(candidates)} aday bulundu. Sonuçları mutlaka doğrulayın.")
                except Exception as exc:
                    st.warning(f"OCR tamamlanamadı: {exc}. Manuel mod kullanılabilir.")
    with tool_col2:
        if st.button("Tüm kayıtları temizle", use_container_width=True):
            st.session_state.records = pd.DataFrame(columns=FORM3_COLUMNS)
            st.session_state.canvas_version += 1
            st.rerun()
    with tool_col3:
        st.caption(
            "OpenAI API anahtarı yoksa, kota/429 hatası oluşursa veya OCR sonuç vermezse uygulama "
            "durmaz; aşağıdaki manuel seçim alanı kullanılabilir."
        )

    page_number = st.selectbox(
        "İşaretlenecek sayfa",
        range(1, len(st.session_state.page_images) + 1),
        format_func=lambda value: f"Sayfa {value}",
    )
    page_image = draw_preview(
        st.session_state.page_images[page_number - 1],
        st.session_state.records,
        page_number,
    )
    original_width, original_height = page_image.size
    display_width = min(1200, original_width)
    display_height = round(original_height * display_width / original_width)
    display_image = page_image.resize((display_width, display_height), Image.Resampling.LANCZOS)

    st.subheader("Manuel balon hedefi seçimi")
    st.caption(
        "Ölçünün üzerine kırmızı nokta koyun. Sonra “Seçilen noktaları balona dönüştür” düğmesine basın. "
        "Balon konumu koordinat tablosundan düzenlenebilir."
    )
    canvas_result = st_canvas(
        fill_color="rgba(220, 20, 45, 0.75)",
        stroke_width=2,
        stroke_color="#DC142D",
        background_image=display_image,
        update_streamlit=True,
        height=display_height,
        width=display_width,
        drawing_mode="point",
        point_display_radius=7,
        display_toolbar=True,
        key=f"canvas_{page_number}_{st.session_state.canvas_version}",
    )

    if st.button("Seçilen noktaları balona dönüştür", type="primary"):
        points = extract_canvas_points(canvas_result.json_data, display_width, display_height)
        if not points:
            st.warning("Önce teknik resim üzerinde en az bir noktaya tıklayın.")
        else:
            current = st.session_state.records.copy()
            next_number = int(max([as_number(value) or 0 for value in current.get("Characteristic No / Balon No", [])], default=0)) + 1
            new_rows = [empty_record(next_number + offset, page_number, x, y) for offset, (x, y) in enumerate(points)]
            st.session_state.records = pd.concat(
                [current, pd.DataFrame(new_rows, columns=FORM3_COLUMNS)],
                ignore_index=True,
            )
            st.session_state.canvas_version += 1
            st.rerun()

    st.subheader("AS9102 Form 3 ölçüm listesi")
    edited = st.data_editor(
        st.session_state.records,
        column_order=FORM3_COLUMNS,
        column_config=editor_configuration(),
        hide_index=True,
        num_rows="dynamic",
        use_container_width=True,
        key="form3_editor",
    )
    edited = apply_suggestions(edited)
    st.session_state.records = edited

    with st.expander("Balon ve hedef koordinatlarını düzenle"):
        if edited.empty:
            st.caption("Henüz kayıt yok.")
        else:
            coordinate_columns = [
                "Characteristic No / Balon No",
                "Sheet",
                "_target_x",
                "_target_y",
                "_balloon_x",
                "_balloon_y",
            ]
            coordinates = st.data_editor(
                edited[coordinate_columns],
                hide_index=True,
                use_container_width=True,
                column_config={
                    "_target_x": st.column_config.NumberColumn("Hedef X (0-1)", min_value=0.0, max_value=1.0, format="%.4f"),
                    "_target_y": st.column_config.NumberColumn("Hedef Y (0-1)", min_value=0.0, max_value=1.0, format="%.4f"),
                    "_balloon_x": st.column_config.NumberColumn("Balon X (0-1)", min_value=0.0, max_value=1.0, format="%.4f"),
                    "_balloon_y": st.column_config.NumberColumn("Balon Y (0-1)", min_value=0.0, max_value=1.0, format="%.4f"),
                },
                key="coordinate_editor",
            )
            for column in coordinate_columns[2:]:
                st.session_state.records[column] = coordinates[column].values

    st.subheader("Form 1 ve Form 2 bilgileri")
    form1_col, form2_col = st.columns(2)
    with form1_col:
        part_info = {
            "part_number": st.text_input("Parça Numarası"),
            "part_name": st.text_input("Parça Adı"),
            "serial_lot": st.text_input("Seri / Lot Numarası"),
            "fai_report_no": st.text_input("FAI Rapor Numarası"),
            "drawing_number": st.text_input("Çizim Numarası", value=Path(uploaded.name).stem),
            "drawing_revision": st.text_input("Çizim Revizyonu"),
            "supplier": st.text_input("Kuruluş / Tedarikçi"),
            "customer": st.text_input("Müşteri"),
            "order_no": st.text_input("Sipariş / İş Emri"),
            "fai_type": st.selectbox("FAI Türü", ["Full FAI", "Partial FAI"]),
        }
    with form2_col:
        process_info = {
            "material": st.text_input("Malzeme"),
            "material_spec": st.text_input("Malzeme Şartnamesi"),
            "certificate_no": st.text_input("Sertifika Numarası"),
            "special_process": st.text_input("Özel Proses"),
            "process_spec": st.text_input("Proses Şartnamesi"),
            "process_supplier": st.text_input("Proses Tedarikçisi"),
            "functional_test": st.text_input("Fonksiyonel Test"),
            "remarks": st.text_area("Form 2 Açıklama"),
        }

    if edited.empty:
        st.warning("PDF ve Excel çıktısı için en az bir balon kaydı ekleyin.")
        return

    duplicate_numbers = edited["Characteristic No / Balon No"].dropna().duplicated().any()
    invalid_coordinates = edited[["_target_x", "_target_y", "_balloon_x", "_balloon_y"]].isna().any().any()
    if duplicate_numbers:
        st.error("Characteristic No / Balon No değerleri benzersiz olmalıdır.")
        return
    if invalid_coordinates:
        st.error("Balon veya hedef koordinatı eksik.")
        return

    try:
        simple_pdf = create_simple_balloon_pdf(pdf_bytes, edited)
        excel_bytes = create_as9102_workbook(edited, part_info, process_info)
    except Exception as exc:
        st.error(f"Çıktılar hazırlanamadı: {exc}")
        return

    st.subheader("Çıktılar")
    pdf_col, excel_col = st.columns(2)
    with pdf_col:
        st.download_button(
            "Sade balonlu PDF indir",
            data=simple_pdf,
            file_name="balonlu_teknik_resim_sade.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    with excel_col:
        st.download_button(
            "AS9102 FAI ölçüm raporunu indir",
            data=excel_bytes,
            file_name="AS9102_FAI_olcum_raporu.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.caption(
        "Sonuç alanları bilinçli olarak doldurulmaz: Ölçüm bekleniyor / Bekliyor / Doldurulacak "
        "değerleri kullanılır. AI veya OCR sonuçları üretim onayı değildir."
    )


if __name__ == "__main__":
    main()
